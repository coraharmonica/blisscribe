# -*- coding: utf-8 -*-
"""
PARSE_LEXICA:

    Contains a class for parsing a language-to-Blissymbols
    dictionary from plaintext (.txt) and Excel (.xlsx) files.

    Allows ease of changing Bliss dictionaries, including
    adding languages beyond English.

    Throughout Blisscribe, "lemma" is meant to be the dictionary
    entry of a word, while "lemma" is meant to be any form of a
    word.  All lemmas are lexemes, but only 1 in a set of lexemes
    is chosen as the lemma.
    e.g. dog == lemma (and == lemma)
         dogs == lemma (and != lemma)

    Alphabetical list of part-of-speech tags used in the
    Penn Treebank Project:

    Number  Tag     Description
    1.      CC      Coordinating conjunction
    2.      CD	    Cardinal number
    3.	    DT	    Determiner
    4.	    EX	    Existential there
    5.	    FW	    Foreign word
    6.	    IN	    Preposition or subordinating conjunction
    7.	    JJ	    Adjective
    8.	    JJR	    Adjective, comparative
    9.	    JJS	    Adjective, superlative
    10.	    LS	    List item marker
    11.	    MD	    Modal
    12.	    NN	    Noun, singular or mass
    13.	    NNS	    Noun, plural
    14.	    NNP	    Proper noun, singular
    15.	    NNPS	Proper noun, plural
    16.	    PDT	    Predeterminer
    17.	    POS	    Possessive ending
    18.	    PRP	    Personal pronoun
    19.	    PRP$	Possessive pronoun
    20.	    RB	    Adverb
    21.	    RBR	    Adverb, comparative
    22.	    RBS	    Adverb, superlative
    23.	    RP	    Particle
    24.	    SYM	    Symbol
    25.	    TO	    to
    26.	    UH	    Interjection
    27.	    VB	    Verb, base form
    28.	    VBD	    Verb, past tense
    29.	    VBG	    Verb, gerund or present participle
    30.	    VBN	    Verb, past participle
    31.	    VBP	    Verb, non-3rd person singular present
    32.	    VBZ	    Verb, 3rd person singular present
    33.	    WDT	    Wh-determiner
    34.	    WP	    Wh-pronoun
    35.	    WP$	    Possessive wh-pronoun
    36.	    WRB	    Wh-adverb
"""
import os
import json
from openpyxl import load_workbook
import re
import blissymbols
from blissymbols import Blissymbol, IMG_PATH, NEW_BLISSYMBOLS

LAST_BLISS_ENCODING = blissymbols.LAST_BLISS_ENCODING
BLISS_SUPPORTED_LANGUAGES = {"English",
                             "Swedish",
                             "Norwegian",
                             "Finnish",
                             "Hungarian",
                             "German",
                             "Dutch",
                             "Afrikaans",
                             "Russian",
                             "Latvian",
                             "Polish",
                             "French",
                             "Spanish",
                             "Portuguese",
                             "Italian",
                             "Danish"}


class LexiconParser:
    FILE_PATH = os.path.dirname(os.path.realpath(__file__))
    LEXICA_PATH = FILE_PATH + "/resources/lexica/"
    LEXICON_PATH = LEXICA_PATH + "universal bliss lexicon.xlsx"
    LEXICON_COLS = ["BCI-AV#",
                    "English",
                    "POS",
                    "Derivation - explanation",
                    "BCI-AV#",
                    "Swedish",
                    "BCI-AV#",
                    "Norwegian",
                    "BCI-AV#",
                    "Finnish",
                    "BCI-AV#",
                    "Hungarian",
                    "BCI-AV#",
                    "German",
                    "BCI-AV#",
                    "Dutch",
                    "BCI-AV#",
                    "Afrikaans",
                    "BCI-AV#",
                    "Russian",
                    "BCI-AV#",
                    "Latvian",
                    "BCI-AV#",
                    "Polish",
                    "BCI-AV#",
                    "French",
                    "BCI-AV#",
                    "Spanish",
                    "BCI-AV#",
                    "Portuguese",
                    "BCI-AV#",
                    "Italian",
                    "BCI-AV#",
                    "Danish"]

    def __init__(self, translator):
        self.translator = translator

    # BLISS DICTS
    # ===========
    def dump_json(self, data, filename):
        """
        Dumps data (prettily) to filename.json.

        :param data: X, data to dump to JSON
        :param filename: str, name of .json file to dump to
        :return: None
        """
        path = self.FILE_PATH + "/" + filename + ".json"
        json.dump(data, open(path, 'w'), indent=4, sort_keys=True)

    def fetch_json(self, filename):
        """
        Returns the official Blissymbols lexicon.

        :param filename: str, name of .json file to fetch
        :return: X, content of given .json file
        """
        lexicon = json.load(open(self.translator.PATH + "/" + filename + ".json"))
        return lexicon

    def fetch_bliss_lexicon(self):
        """
        Returns the official Blissymbols lexicon.

        :return: dict, where...
            key (str) - Blissymbol word (in English)
            val (dict) - dict corresponding to Blissymbol word
        """
        return self.fetch_json("bliss_lexicon")

    def refresh_bliss_lexicon(self):
        """
        Refreshes the Blissymbols JSON lexicon to include all new entries in this
        LexiconParser's bliss_dict.

        :return: None
        """
        new_lexicon = dict()
        bliss_dict = self.translator.get_eng_bliss_dict()

        for entry in bliss_dict:
            new_lexicon.setdefault(entry, list())
            defns = bliss_dict[entry]
            for blissymbol in defns:
                bd = self.blissymbol_to_dict(blissymbol)
                if bd is not None:
                    new_lexicon[entry].append(bd)

        self.dump_json(new_lexicon, "bliss_lexicon")

    def refresh_bliss_encoding(self):
        """
        Overwrites the source Blissymbols-to-unicode encoding
        dictionary with this LexiconParser's BlissTranslator's
        bliss_to_unicode dict.

        :return: None
        """
        encoding = self.translator.get_bliss_to_unicode()
        self.dump_json(encoding, "bliss_encoding")

    def refresh_bliss_decoding(self):
        """
        Overwrites the source unicode-to-Blissymbols decoding
        dictionary with this LexiconParser's BlissTranslator's
        unicode_to_bliss dict.

        :return: None
        """
        decoding = self.translator.get_unicode_to_bliss()
        self.dump_json(decoding, "bliss_decoding")

    def init_bliss_lexicon(self, language):
        """
        Initializes a Blissymbols lexicon in the given language.

        :param language: str, desired Blissymbol lexicon language
        :return: dict, where...
            key (str) - words in given language
            val (Set(Blissymbol)) - corresponding Blissymbols with
                translations in given language and English
        """
        bliss_dict = {}
        lexicon = self.fetch_bliss_lexicon()

        for entry in lexicon:
            bliss_words = lexicon[entry]
            for bliss_word in bliss_words:
                blissymbol = self.dict_to_blissymbol(bliss_word)
                if blissymbol is not None:
                    if language != "English":
                        lang_words = blissymbol.get_translation(language)
                        for lang_word in lang_words:
                            bliss_dict.setdefault(lang_word, set())
                            bliss_dict[lang_word].add(blissymbol)
                    else:
                        bliss_dict.setdefault(entry, set())
                        bliss_dict[entry].add(blissymbol)

        return bliss_dict

    def init_bliss_encoding(self):
        """
        Returns a Blissymbols-to-unicode encoding dictionary.

        :return: dict, where...
            key (str) - Blissymbol word (in English)
            val (List[str]) - unicode representations of Blissymbol
        """
        return self.fetch_json("bliss_encoding")

    def init_bliss_decoding(self):
        """
        Returns a unicode-to-Blissymbols decoding dictionary.

        :return: dict, where...
            key (str) - unicode representation of a Blissymbol
            val (List[str]) - words corresponding to given unicode
        """
        return self.fetch_json("bliss_decoding")

    # MULTILINGUAL
    # ============
    def parse_lexicon(self, language):
        """
        Parses plaintext file for given language.
        Returns a dict with all words in lexicon
        as keys and corresponding lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes first word on every line is the lemma form
        of all subsequent words on the same line.
        ~
        N.B. The same lemma value will often belong to multiple keys.

        e.g. "kota, kocie, kot" -> {"kota":"kota", "kocie":"kota", "kot":"kota"}

        :param language: str, language of .txt file for lexicon
        :return: dict, where...
            key (str) - inflected form of a word
            val (str) - lemma form of inflected word
        """
        filename = "/resources/lexica/" + language + ".txt"
        lexicon = None

        with open(self.FILE_PATH + filename, "rb") as lex:
            if language == "Polish":
                lexicon = self.parse_polish_lexicon(lex)
            elif language == "French":
                lexicon = self.parse_french_lexicon(lex)

        return lexicon

    def parse_french_lexicon(self, lex):
        """
        Parses plaintext file for French lexicon with
        given filename.  Returns a dict with all lexemes
        in French lexicon as keys and corresponding
        lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes second word on every line is the lemma form
        of previous words on the same line.
        ~
        N.B. The same lemma value will often belong to
        multiple lexemes.

        e.g. "grande, grand" -> {"grande":"grand", "grand":"grand"}

        :param lex: List[str], entries in French lexicon
        :return: dict, where...
            key (str) - any lexical form of a word
            val (str) - lemmatized form of lemma
        """
        lexicon = {}

        for line in lex:
            line = line.decode("utf-8")
            line = line.strip("\n")
            if line[-1] == "=":
                lemma = line[:-2]
                lexicon[lemma] = lemma
            else:
                entry = line.split("\t")
                lemma = entry[1]
                lexeme = entry[0]
                lexicon[lexeme] = lemma

        return lexicon

    def parse_polish_lexicon(self, lex):
        """
        Parses plaintext file for Polish lexicon with
        given filename.  Returns a dict with all lexemes
        in Polish lexicon as keys and corresponding
        lemma forms as values.
        ~
        Inflected forms in each lexical entry should be
        separated by ",".
        ~
        Assumes first word on every line is the lemma form
        of following words on the same line.
        ~
        N.B. The same lemma value will often belong to
        multiple lexemes.

        e.g. "kota, kot, kocie" -> {"kota":"kota", "kot":"kota", "kocie":"kota"}

        :param lex: List[str], entries in Polish lexicon
        :return: dict, where...
            key (str) - any lexical form of a word
            val (str) - lemmatized form of lemma
        """
        lexicon = {}

        for line in lex:
            line = line.decode("utf-8")
            line = line.strip("\n")
            line = line.strip("\r")
            inflexions = line.split(",")
            lemma = inflexions[0]

            for inflexion in inflexions:
                lexicon[inflexion.strip()] = lemma

        return lexicon

    # ENCODINGS
    # =========
    def get_bliss_encoding(self, filename=None):
        """
        Reads plaintext file with given filename and returns
        a list of Blissymbol names.
        ~
        Output conforms to encoding suggested here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf

        :param filename: str, .txt file with Blissymbol names
        :return: List[str], Blissymbol names
        """
        if filename is None:
            filename = self.LEXICA_PATH + "blissymbol encoding.txt"
        else:
            filename = self.FILE_PATH + filename

        defns = []

        with open(filename, "rb") as bliss_in:
            for bliss_name in bliss_in:
                name = bliss_name[11:-1].lower()
                defns.append(name)

        return defns

    def write_bliss_encoding(self, unicode_keys=False):
        """
        Writes and returns a dict linking Blissymbol names to unicode values,
        or, if unicode_keys=True, linking unicode keys to Blissymbol names.
        ~
        If unicode_keys is set to True, this method names the file's output
        dictionary UNICODE_TO_BLISS.  Otherwise, its name is BLISS_TO_UNICODE.
        ~
        Output conforms to encoding suggested here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf
        ~
        Used to update BLISS_TO_UNICODE and UNICODE_TO_BLISS as needed.

        :param unicode_keys: bool, whether output dict keys are unicode
        :return: dict, where...
            key (str) - word definition (unicode if unicode_keys=True)
            val (str) - unicode (word definition if unicode_keys=True)
        """
        defns = self.get_bliss_encoding()
        encodings = {}

        if unicode_keys:
            title = "UNICODE_TO_BLISS"
        else:
            title = "BLISS_TO_UNICODE"

        write_filename = self.FILE_PATH + "/bliss_encoding.py"
        idx = "3200"

        with open(write_filename, "a") as bliss_out:
            bliss_out.write(title + " = {")

            for defn in defns:
                uni = "U+" + idx  # make unicode key

                if unicode_keys:
                    key = uni
                    val = defn
                else:
                    key = defn
                    val = uni

                encoding = '\n    "' + key + '": ["' + val + '"],'
                bliss_out.write(encoding)
                encodings[key] = [val]

                idx = int(idx, 16) + 1
                idx = hex(idx)[2:]

            bliss_out.write("\n    }\n\n")

        global LAST_BLISS_ENCODING
        LAST_BLISS_ENCODING = idx

        return encodings

    # ILI
    # ===
    def read_ili_mapping(self):
        """
        Reads plaintext file with mapping from Princeton WordNet
        to ILI (Interlingual Language Index), a conceptual dictionary.
        ~
        Used for cross-lingual translation.

        :return: List[ILIEntry], list of ILI definitions
        """
        defns = []
        with open(self.LEXICA_PATH + "ili-wn-mapping.txt", "r") as ili:
            lines = ili.readlines()[10:]
            for defn in lines:
                defns.append(self.clean_ili_defn(defn))
        return defns

    def write_ili_mapping(self, clean_defns):
        """
        Writes to plaintext file with mapping from Princeton WordNet
        to ILI (Interlingual Language Index), a conceptual dictionary.
        ~
        Used for cross-lingual translation.

        :param clean_defns: List[ILIEntry], list of word entries to write
        :return: None
        """
        out = self.LEXICA_PATH + "ili-wn-mapping-cleaned.txt"
        open(out, 'w').close()  # wipe file before writing

        with open(out, "a") as ili:
            for defn in clean_defns:
                ili.write(str(defn) + "\n")

    def read_write_ili_mapping(self):
        """
        Reads ILI mapping and writes cleaned definitions
        to file.

        :return: None
        """
        clean_defns = self.read_ili_mapping()
        self.write_ili_mapping(clean_defns)

    def clean_ili_defn(self, defn):
        """
        Cleans the input ILI definition line.

        :param defn: str, line in ILI
        :return: ILIEntry, an entry for a single ILI concept
        """
        idx = re.search(pattern="i[0-9]{1,8}", string=defn).group(0)
        locn = re.search(pattern="[0-9]{8}-[a|n|r|s|v]", string=defn).group(0)
        word = re.search(pattern="#\s.+", string=defn).group(0)
        word = str(word[2:])
        words = word.split(",")
        words = [word.strip() for word in words]
        entry = self.ili_dict.makeEntry(int(idx[1:]), str(locn), words)
        return entry

    # WORDNET
    # =======
    def write_bliss_wordnet(self, bliss_dict):
        """
        Writes mapping between Blissymbol words and
        Synsets to plaintext file.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol word in English
            val (List[Synset]) - synsets associated with Blissymbol
        :return: None
        """
        blissnet = self.bliss_dict_to_encoding_wordnet(bliss_dict)
        self.write_bliss_wordnet_encoding(blissnet)

    def write_bliss_wordnet_encoding(self, blissnet):
        """
        Writes mapping between Blissymbol unicodes and
        Synsets to plaintext file.

        :param blissnet: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        :return: None
        """
        out = self.LEXICA_PATH + "blissnet.txt"

        with open(out, "w") as wordnet:
            wordnet.write(self.encoding_dict_to_str(blissnet))
            wordnet.close()

    def bliss_dict_to_wordnet(self, bliss_dict):
        """
        Returns a dictionary of Blissymbol word keys and synsets
        from the given bliss_dict.
        ~
        N.B. Output will be most accurate with a multilingual
        Bliss dictionary.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol word in English
            val (List[Blissymbol]) - corresponding Blissymbols with
                translations in all languages
        :return: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        """
        wordnet = {}

        for word in bliss_dict:
            for blissymbol in bliss_dict[word]:
                uni = blissymbol.get_unicode()
                synsets = blissymbol.get_synsets()
                wordnet[uni] = synsets

        return wordnet

    def bliss_dict_to_encoding_wordnet(self, bliss_dict):
        """
        Returns a dictionary of Bliss unicodes and synsets
        from the given bliss_dict.
        ~
        N.B. Output will be most accurate with a multilingual
        Bliss dictionary.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol word in English
            val (List[Blissymbol]) - corresponding Blissymbols with
                translations in all languages
        :return: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        """
        wordnet = {}

        for word in bliss_dict:
            for blissymbol in bliss_dict[word]:
                uni = blissymbol.get_unicode()
                synsets = blissymbol.get_synsets()
                wordnet[uni] = synsets

        return wordnet

    def get_tab_file(self, lang_code):
        """
        Retrieves a multilingual WordNet tab file for the
        given language.
        ~
        If no such tab file exists, returns None.

        :param lang_code: str, 3-character ISO language code
        :return: tab file, WordNet file for given language
        """
        try:
            return open(self.FILE_PATH + "/resources/omw_tabs/" +
                        "wn-cldr-" + lang_code + ".tab")
        except Exception:
            return None

    def make_blissymbol(self):
        """
        Prompts user for new Blissymbol entry information.

        :return: Blissymbol, represents 1 Bliss lexical entry
        """
        bliss_name = raw_input("What do you call your new Blissymbol? ")
        bliss_filename = bliss_name + ".png"
        print("Which part of speech is this? ")
        code = Blissymbol.POS_COLOUR_CODE

        for pos in code:
            print(str(pos) + ":\t" + str(code[pos]))

        pos = raw_input("")
        print("Which atomic Blissymbols is this made of? ")
        derivations = raw_input("Separate them by commas: ")
        derivations = derivations.split(",")
        derivs = ""
        derivs += "("

        for idx in range(0, len(derivations)):
            derivs += derivations[idx]
            if idx == len(derivations)-1:
                derivs += ")"
            else:
                derivs += " + "

        translations = {}

        for language in BLISS_SUPPORTED_LANGUAGES:
            print("What is/are the translation(s) in " + language + "? ")

            try:
                translation = raw_input("Separate by commas if necessary: ")
            except SyntaxError:
                continue
            else:
                translation = translation

            translations.setdefault(language, [])
            translations[language].append(translation)

        blissymbol = Blissymbol(img_filename=bliss_filename, pos=pos, derivation=derivs,
                                translations=translations, translator=self.translator)
        return blissymbol

    # XLSX ENTRIES
    # ============
    def blissymbol_to_xlsx_entry(self, blissymbol):
        """
        Converts the given Blissymbol to a list of
        information constituting a Bliss lexicon entry.
        ~
        Must be in order of LEXICON_COLS.

        :param blissymbol: Blissymbol, symbol to convert to entry
        :return: List[str], Bliss lexicon entry
        """
        row = []
        bci_col = self.LEXICON_COLS[0]
        pos_col = self.LEXICON_COLS[2]
        deriv_col = self.LEXICON_COLS[3]
        uni = blissymbol.get_unicode() #[0]
        uni = uni[2:]

        for col in self.LEXICON_COLS:
            if col == bci_col:
                row.append("C+" + uni)
            elif col == pos_col:
                row.append(blissymbol.pos_to_int())
            elif col == deriv_col:
                row.append(blissymbol.get_derivation())
            else:
                translations = blissymbol.get_translation(col)
                translations = u",".join(translations)
                translations = self.translator.deunicodize(translations)
                row.append(translations)

        return row

    def append_bliss_xlsx_entry(self, blissymbol=None):
        """
        Adds data to bliss lexicon file for later use.
        ~
        If no Blissymbol is provided, prompts you to create your
        own custom symbol.

        :param blissymbol: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        if not blissymbol:
            blissymbol = self.make_blissymbol()

        print("making new Blissymbol entry for " + str(blissymbol))
        bliss_entry = self.blissymbol_to_xlsx_entry(blissymbol)

        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        sheet.append(bliss_entry)

        book.save(self.LEXICON_PATH)

    def extend_bliss_xlsx_entry(self, blissymbol):
        """
        Extends data in bliss lexicon file for later use.

        :param bliss_entry: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        print("extending entry for ", blissymbol.get_bliss_name())
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        start_row = sheet.min_row
        end_row = sheet.max_row

        for row_idx in range(start_row, end_row):
            cell = sheet.cell(row=row_idx, column=2)
            eng_word = cell.value
            eng_words = eng_word.split(",")
            for synonym in eng_words:
                synonym = synonym.strip()
                if synonym == blissymbol.get_bliss_name():
                    print("found English Blissymbol synonym: ", eng_word)
                    translations = blissymbol.get_translations()
                    for language in translations:
                        defns = translations[language]
                        if len(defns) != 0:
                            lang_idx = self.LEXICON_COLS.index(language)
                            edit = sheet.cell(row=row_idx, column=lang_idx+1)

                            if edit.value is None:
                                edit.value = ""
                            else:
                                edit.value = self.translator.unicodize(edit.value)
                            values = []
                            for defn in defns:
                                values = edit.value.split(u",")
                                if defn not in values:
                                    values.append(defn)
                            values = self.translator.remove_duplicates(values)
                            edit.value = ",".join(values)
                            print(u"edited " + language + u" to be " + edit.value)
                    break

        book.save(self.LEXICON_PATH)

    # XLSX PARSING
    # ============
    def get_img_filenames(self):
        """
        Reads the given XLS file for a cross-lingual Bliss
        dictionary and returns a list of image filenames.

        :param filename: str, name of XLS file
        :return: List[str], image filenames
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        img_col = 2

        for col in sheet.iter_cols(min_col=img_col, max_col=img_col):
            imgs = [str(row.value) + ".png" for row in col[1:]]
            return imgs

    def get_defns(self, language):
        """
        Returns a list of the given language's words
        in the XLSX Bliss lexicon.

        :param language: str, output list's desired language
        :return: List[str], given language's Bliss words
        """
        assert language in BLISS_SUPPORTED_LANGUAGES

        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]

        defns = []

        for col in sheet.iter_cols():
            if col[0].value == language:
                defns = [row.value for row in col[1:]]
                break

        if len(defns) == 0:
            raise IOError(language + " is not in Blissymbolics lexicon.")

        return defns

    def get_parts_of_speech(self):
        """
        Returns an ordered list of all parts of speech
        from the XLSX Bliss lexicon.

        :return: List[str], list of parts of speech
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        pos_col = 3

        for col in sheet.iter_cols(min_col=pos_col, max_col=pos_col):
            pos = [row.value for row in col[1:]]
            return pos

    def get_derivations(self):
        """
        Returns an ordered list of derivations from
        the XLSX Bliss lexicon.

        :return: List[str], list of derivations
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        deriv_col = 4

        for col in sheet.iter_cols(min_col=deriv_col, max_col=deriv_col):
            derivations = [unicode.encode(row.value, 'ascii', errors='ignore') for row in col[1:]]
            return derivations

    def get_defn_words(self, defn):
        """
        Separates given defn by commas and returns
        the resulting list.
        ~
        This method ensures that output meets
        all criteria for being added to Blisscribe.

        :param defn: str, Blissymbol definition
        :return: List[str], defn's words separated by commas
        """
        words = []
        if defn is not None:
            if not self.is_old(defn):  # no outdated definitions
                for word in defn.split(","):
                    word = self.parse_alphabetic(word)
                    if not self.is_indicator(word):
                        word = self.strip_parens(word)
                    if len(word) != 0:
                        words.append(word)
        return words

    # HELPERS
    # =======
    def is_indicator(self, word):
        """
        Returns True if this Blissymbol word is an indicator.

        :param word: str, Blissymbol word from bliss_dict
        :return: bool, True if given word is an indicator
        """
        return word[:9] == "indicator"

    def is_old(self, word):
        """
        Returns True if this Blissymbol word is marked OLD.

        :param word: str, Blissymbol word from bliss_dict
        :return: bool, True if given word is old
        """
        return word[-5:] == "(OLD)"

    def is_letter(self, img_filename):
        """
        Returns True if given image filename suggests that its
        corresponding Blissymbol is a letter of the alphabet.
        ~
        Checks if img_filename ends with "ercase" (for uppercase
        or lowercase).
        ~
        Used for excluding alphabetic-only Bliss characters.

        :param img_filename: str, image filename of a Blissymbol
        :return: bool, whether given image is an alphabetic letter
        """
        return img_filename[-11:] == "ercase).png"

    def strip_parens(self, word):
        """
        Strips parenthetical(s) from the given word.
        ~
        String cuts off just before first set of parentheses.

        e.g. strip_parens("English_(language)") -> "English"

        :param word: str, word to strip parentheticals from
        :return: str, word with parentheticals stripped
        """
        new_word = []
        remove = False

        for char in word:
            if remove == False and char != "(":
                new_word.append(char)
            elif char == "(":
                remove = True
            elif char == ")":
                remove = False

        new_word = ("".join(new_word)).strip()
        return new_word

    def parse_alphabetic(self, word):
        """
        Parses the given non-alphabetic word into an
        alphabetic-only version of the word.

        e.g. parse_alphabetic("English_(language)") -> "English (language)"

        :param word: str, non-alphabetic word
        :return: str, alphabetic version of input word
        """
        word = word.replace("_", " ")
        word = word.replace("-", " ")
        word = word.replace("!", "")
        return word

    # BLISS MANIPULATION
    # ==================
    def blissymbol_to_str(self, blissymbol):
        """
        Returns the given Blissymbol formatted as a string for
        the bliss_lexicon file.

        :param blissymbol: Blissymbol, blissymbol to turn to string
        :return: (unicode) str, blissymbol as string
        """
        return self.translator.unicodize('("' +
                                         blissymbol.get_bliss_name() + '", "' +
                                         blissymbol.get_pos() + '", "' +
                                         blissymbol.get_derivation() + '", ' +
                                         self.dict_to_str(blissymbol.get_translations()) +
                                         ')')

    def blissymbol_to_dict(self, blissymbol):
        """
        Returns a dict of this Blissymbol's initializing
        parameters, i.e. its name, pos, derivation, and translations.

        :param blissymbol: Blissymbol, symbol to turn to tuple
        :return: dict, where...
            key (str) - name of Blissymbol field
            val (X) - corresponding value
        """
        return {u"name": blissymbol.get_bliss_name(),
                u"pos": blissymbol.get_pos(),
                u"derivation": blissymbol.get_derivation(),
                u"translations": blissymbol.get_translations()}

    # DICT MANIPULATION
    # =================
    def bliss_dict_to_str(self, bliss_dict):
        """
        Returns the given dictionary with each of its keys and values
        as strings.

        :param bliss_dict: dict, input dictionary to turn to string
        :return: (unicode) str, input dictionary turned to string
        """
        res = []

        for key in sorted(bliss_dict.keys()):
            val = bliss_dict[key]
            if len(val) != 0:
                res.append(u'\t"' + self.translator.unicodize(key) + u'": [')
                for blissymbol in val:
                    synsets = sorted(blissymbol.get_synsets())
                    synsets = [str(synset) for synset in synsets]
                    #items = sorted([str(blissymbol.get_synset()) for blissymbol in val])
                    #items = [self.blissymbol_to_str(item) for item in val]
                    res.append(u',\n\t\t'.join(synsets))
                res.append(u"],\n")

        return u"".join(res)

    def bliss_dict_to_str_dict(self, bliss_dict):
        """
        Returns the given dictionary as the string it would be
        written as in Python.

        :param bliss_dict: dict, input dictionary to turn to string
        :return: (unicode) str, input dictionary turned to string
        """
        res = [u"BLISS_LEXICON = {\n"]
        res.append(self.bliss_dict_to_str(bliss_dict))
        res.append(u"}")
        return u"".join(res)

    def dict_to_blissymbol(self, d):
        """
        Returns the given dict, d, as a Blissymbol.
        ~
        This method assumes d has keys named
        "name", "pos", "derivation", and "translations",
        and initializes a Blissymbol with its fields corresponding
        to these keys.

        :param d: dict, dictionary to turn into Blissymbol
        :return: Blissymbol, d as a Blissymbol
        """
        name = d[u"name"]
        pos = d[u"pos"]

        try:
            derivation = d[u"derivation"]
        except KeyError:
            derivation = u""
        try:
            translations = d[u"translations"]
        except KeyError:
            translations = {}

        blissymbol = Blissymbol(name+u".png",
                                pos,
                                derivation,
                                translations,
                                self.translator)
        return blissymbol

    def print_dict(self, d):
        """
        Prints the given dictionary as if it were written in code.
        Used for visualizing and copying dict contents.

        :param d: dict, input dictionary to print
        :return: None
        """
        print(self.encoding_dict_to_str(d))