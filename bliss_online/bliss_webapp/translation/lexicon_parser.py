# -*- coding: utf-8 -*-
"""
PARSE_LEXICA:

    Contains a class for parsing a language-to-Blissymbols
    dictionary from plaintext (.txt) and Excel (.xlsx) files.

    Allows ease of changing Bliss dictionaries, including
    adding languages beyond English.

    Throughout Blisscribe, "lemma" is meant to be the dictionary
    entry of a word, while "lemma" is meant to be any form of a
    word.  All lemmas are lexemes, but only 1 in a set of lexemes
    is chosen as the lemma.
    e.g. dog == lemma (and == lemma)
         dogs == lemma (and != lemma)

    Alphabetical list of part-of-speech tags used in the
    Penn Treebank Project:

    Number  Tag     Description
    1.      CC      Coordinating conjunction
    2.      CD	    Cardinal number
    3.	    DT	    Determiner
    4.	    EX	    Existential there
    5.	    FW	    Foreign word
    6.	    IN	    Preposition or subordinating conjunction
    7.	    JJ	    Adjective
    8.	    JJR	    Adjective, comparative
    9.	    JJS	    Adjective, superlative
    10.	    LS	    List item marker
    11.	    MD	    Modal
    12.	    NN	    Noun, singular or mass
    13.	    NNS	    Noun, plural
    14.	    NNP	    Proper noun, singular
    15.	    NNPS	Proper noun, plural
    16.	    PDT	    Predeterminer
    17.	    POS	    Possessive ending
    18.	    PRP	    Personal pronoun
    19.	    PRP$	Possessive pronoun
    20.	    RB	    Adverb
    21.	    RBR	    Adverb, comparative
    22.	    RBS	    Adverb, superlative
    23.	    RP	    Particle
    24.	    SYM	    Symbol
    25.	    TO	    to
    26.	    UH	    Interjection
    27.	    VB	    Verb, base form
    28.	    VBD	    Verb, past tense
    29.	    VBG	    Verb, gerund or present participle
    30.	    VBN	    Verb, past participle
    31.	    VBP	    Verb, non-3rd person singular present
    32.	    VBZ	    Verb, 3rd person singular present
    33.	    WDT	    Wh-determiner
    34.	    WP	    Wh-pronoun
    35.	    WP$	    Possessive wh-pronoun
    36.	    WRB	    Wh-adverb
"""
import os
import json
from openpyxl import load_workbook
from blissymbols import Blissymbol, NEW_BLISSYMBOLS


class LexiconParser:
    FILE_PATH = os.path.dirname(os.path.realpath(__file__))
    RESOURCE_PATH = FILE_PATH + "/resources/"
    DATA_PATH = RESOURCE_PATH + "data/"
    WORDNET_PATH = RESOURCE_PATH + "wordnet/"
    LEXICA_PATH = RESOURCE_PATH + "lexica/"
    LEXICON_PATH = LEXICA_PATH + "universal bliss lexicon.xlsx"
    LEXICON_COLS = ["BCI-AV#",
                    "English",
                    "POS",
                    "Derivation - explanation",
                    "BCI-AV#",
                    "Swedish",
                    "BCI-AV#",
                    "Norwegian",
                    "BCI-AV#",
                    "Finnish",
                    "BCI-AV#",
                    "Hungarian",
                    "BCI-AV#",
                    "German",
                    "BCI-AV#",
                    "Dutch",
                    "BCI-AV#",
                    "Afrikaans",
                    "BCI-AV#",
                    "Russian",
                    "BCI-AV#",
                    "Latvian",
                    "BCI-AV#",
                    "Polish",
                    "BCI-AV#",
                    "French",
                    "BCI-AV#",
                    "Spanish",
                    "BCI-AV#",
                    "Portuguese",
                    "BCI-AV#",
                    "Italian",
                    "BCI-AV#",
                    "Danish"]
    BLISS_LANGS = {"English",
                   "Swedish",
                   "Norwegian",
                   "Finnish",
                   "Hungarian",
                   "German",
                   "Dutch",
                   "Afrikaans",
                   "Russian",
                   "Latvian",
                   "Polish",
                   "French",
                   "Spanish",
                   "Portuguese",
                   "Italian",
                   "Danish"}

    def __init__(self, translator):
        self.translator = translator
        self.blissymbols = None
        self.derivations = None

    # JSON
    # ====
    def dump_json(self, data, filename):
        """
        Dumps data (prettily) to filename JSON.

        :param data: X, data to dump to JSON
        :param filename: str, name of .json file to dump to
        :return: None
        """
        path = self.DATA_PATH + "/" + filename + ".json"
        json.dump(data, open(path, 'w', encoding='utf-8'), indent=1, sort_keys=True, ensure_ascii=False)

    def load_json(self, filename):
        """
        Returns the official Blissymbols lexicon.

        :param filename: str, name of .json file to fetch
        :return: X, content of given .json file
        """
        return json.load(open(self.DATA_PATH + filename + ".json", encoding='utf-8'))

    # BLISS DATA
    # ==========
    #   1)  All Blissymbols
    #   2)  BCI-Blissname Map
    #   3)  BCI Blissnet
    #   4)  Bliss Derivations
    #   5)  Bliss Unicode En-/Decoding
    #   6)  Bliss Lexicon
    #   7)  Bliss Unicode
    #   8)  Blissnet
    #   9)  Blisswordnet
    #   10) WordNet 3.0-to-3.1 Map

    # ALL BLISSYMBOLS
    # ---------------
    def load_all_blissymbols(self):
        """
        Returns a list of dictionaries, each of which represents a
        Blissymbol from the official Blissymbols lexicon.

        :return: List[dict(str, X)], where...
            key (str) - Blissymbol's attribute
            val (X) - attribute's value
        """
        return self.load_json("all_blissymbols")

    def all_blissymbols(self):
        """
        Creates and returns a list of dictionaries for all Blissymbols.

        :param language: str, desired Blissymbol lexicon language
        :return: List[Blissymbol], all Blissymbols from Blissymbol lexicon
        """
        all_blissymbols = list()
        blissymbols = self.load_all_blissymbols()

        for bs in blissymbols:
            blissymbol = self.dict_to_blissymbol(bs)
            if blissymbol is not None:
                all_blissymbols.append(blissymbol)

        return all_blissymbols

    def init_blissymbols(self):
        """
        Returns a list of all Blissymbols.

        :param language: str, desired Blissymbol lexicon language
        :return: List[Blissymbol], all Blissymbols from Blissymbol lexicon
        """
        if self.blissymbols is None:
            self.blissymbols = self.all_blissymbols()
        return self.blissymbols

    # BCI-BLISSNAME MAP
    # -----------------
    def load_bci_blissname_map(self):
        """
        Returns a dict from BCI-AV ID numbers to their
        corresponding Blissymbol's name.
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict, where...
            key (int) - BCI-AV number for a Blissymbol
            val (str) - Blissymbol's name
        """
        return self.load_json("bci-av_mapping")

    def fresh_bci_blissname_map(self):
        """
        Returns a dict mapping BCI-AV numbers to Blissymbol
        names and dumps it to bci-av_map_blissnames JSON.
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict(int, str), where...
            key (int) - BCI-AV number for a Blissymbol
            val (str) - Blissymbol's name
        """
        bliss_dict = self.load_all_blissymbols()
        bci_mapping = dict()

        for entry in bliss_dict:
            bci_av = int(entry["BCI-AV"])
            name = entry["name"]
            bci_mapping[bci_av] = name

        self.dump_json(bci_mapping, 'bci-av_mapping')
        return bci_mapping

    # BCI BLISSNET
    # ------------
    def load_bci_blissnet(self):
        """
        Returns a dict from BCI-AV ID numbers to their
        corresponding Blissymbol's list of PWN Synset strings.
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict, where...
            key (int) - BCI-AV number for a Blissymbol
            val (List[str]) - WordNet Synset strings
        """
        return self.load_json("bci_blissnet")

    def fresh_bci_blissnet(self):
        """
        Returns a dict mapping BCI-IV Blissymbol IDs
        to a list of WordNet synset strings, and
        dumps it to bci_blissnet JSON.

        :return: dict(int, list), where...
            key (int) - BCI-AV number
            val (List[str]) - strings for WordNet synsets
        """
        bci_map = self.load_bci_blissname_map()
        bliss_map = {bliss: bci for bci, bliss in bci_map.items()}
        blissnet_map = self.load_blisswordnet()
        net_bci_map = dict()

        for blissword in blissnet_map:
            bci = int(bliss_map[blissword])
            net_bci_map[bci] = blissnet_map[blissword]

        self.dump_json(net_bci_map, "bci_blissnet")
        return net_bci_map

    # BLISS DERIVATIONS
    # -----------------
    def load_bliss_derivations(self):
        """
        Returns a dictionary of Blissymbol derivation strings and
        the list of derivations belonging to that string.
        ~
        Each Blissymbol derivation string consists of 2+ names for
        the Blissymbols it is composed of, or is otherwise atomic.

        :return: dict, where...
            key (str) - Blissymbol derivation string
            val (list) - list of Blissymbol names in this derivation
        """
        return self.load_json("bliss_derivations")

    def fresh_bliss_derivations(self):
        """
        Returns a fresh derivations dictionary for translating
        Blissymbol derivations to words.

        :return: dict, where...
            key (str) - unicode for Blissymbol
            val (List[str]) - Blissymbol names for given unicode
        """
        all_blissymbols = self.init_blissymbols()
        bliss_derivations = dict()

        for blissymbol in all_blissymbols:
            bliss_derivations[blissymbol.bliss_name] = blissymbol.derivations

        self.dump_json(bliss_derivations, "bliss_derivations")
        return bliss_derivations

    def init_derivations(self):
        """
        Returns a dictionary of all Blissymbol names and their derivations.

        :return: dict, where...
            key (str) - Blissymbol derivation string
            val (list) - list of Blissymbol names in this derivation
        """
        if self.derivations is None:
            self.derivations = self.load_bliss_derivations()
        return self.derivations

    # BLISS UNICODE EN-/DECODING
    # --------------------------
    def load_bliss_encoding(self):
        """
        Returns a Blissymbols-to-unicode encoding dictionary.
        ~
        Encoding conforms to suggestions here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf

        :return: dict, where...
            key (str) - Blissymbol word (in English)
            val (List[str]) - unicode representations of Blissymbol
        """
        return self.load_json("bliss_encoding")

    def fresh_bliss_encoding(self):
        """
        Returns a fresh text_image-to-unicode dictionary for translating
        words to Blissymbols.

        :return: dict, where...
            key (str) - word for a Blissymbol
            val (List[str]) - unicode Blissymbols for this word
        """
        unicodes = self.load_bliss_unicode()
        encoding = dict()

        for uni in unicodes:
            bliss_entry = unicodes[uni]
            if bliss_entry is not None:
                bliss_str = self.translator.deunderscore(bliss_entry)
                bliss_strs = bliss_str.split(",")
                for bliss in bliss_strs:
                    #bliss = self.translator.remove_parens(bliss).rstrip("-")
                    encoding.setdefault(bliss, list())
                    encoding[bliss].append(uni)

        self.dump_json(encoding, "bliss_encoding")
        return encoding

    def refresh_bliss_encoding(self):
        """
        Overwrites the source Blissymbols-to-unicode encoding
        dictionary with this LexiconParser's BlissTranslator's
        bliss_to_unicode dict.

        :return: None
        """
        encoding = self.translator.get_bliss_to_unicode()
        self.dump_json(encoding, "bliss_encoding")

    def load_bliss_decoding(self):
        """
        Returns a unicode-to-Blissymbols decoding dictionary.
        ~
        Encoding conforms to suggestions here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf

        :return: dict, where...
            key (str) - unicode representation of a Blissymbol
            val (List[str]) - English words corresponding to given unicode
        """
        return self.load_json("bliss_decoding")

    def fresh_bliss_decoding(self):
        """
        Returns a fresh unicode-to-text_image dictionary for translating
        Blissymbols to words.

        :return: dict, where...
            key (str) - unicode for Blissymbol
            val (List[str]) - Blissymbol names for given unicode
        """
        unicodes = self.load_bliss_unicode()
        decoding = dict()

        for uni in unicodes:
            bliss_entry = unicodes[uni]
            if bliss_entry is not None:
                bliss_str = self.translator.deunderscore(bliss_entry)
                bliss_strs = bliss_str.split(",")
                decoding[uni] = bliss_strs

        self.dump_json(decoding, "bliss_decoding")
        return decoding

    def refresh_bliss_decoding(self):
        """
        Overwrites the source unicode-to-Blissymbols decoding
        dictionary with this LexiconParser's BlissTranslator's
        unicode_to_bliss dict.

        :return: None
        """
        decoding = self.translator.get_unicode_to_bliss()
        self.dump_json(decoding, "bliss_decoding")

    # BLISS LEXICON
    # -------------
    def load_bliss_lexicon(self):
        """
        Returns the official Blissymbols lexicon.

        :return: dict, where...
            key (str) - Blissymbol word (in English)
            val (dict) - dict corresponding to Blissymbol word
        """
        return self.load_json("bliss_lexicon")

    def init_bliss_lexicon(self, language):
        """
        Initializes a Blissymbols lexicon in this language.

        :param language: str, desired Blissymbol lexicon language
        :return: dict, where...
            key (str) - word in this language
            val (Set(Blissymbol)) - Blissymbols for word
        """
        bliss_dict = {}
        all_blissymbols = self.all_blissymbols()

        for blissymbol in all_blissymbols:
            lang_words = blissymbol.get_translation(language)

            for lang_word in lang_words:
                bliss_dict.setdefault(lang_word, set())
                bliss_dict[lang_word].add(blissymbol)

        return bliss_dict

    def fresh_bliss_lexicon(self):
        """
        Returns a new Blissymbols lexicon and overwrites
        bliss_lexicon.json with it.

        :return: dict, where...
            key (str) - word for Blissymbol
            val (List[dict]) - Blissymbol dictionaries for this word
        """
        all_blissymbols = self.init_blissymbols()
        bci_map = self.load_json("bci_blissname_map")
        bci_map = {v: k for k, v in bci_map.items()}
        bliss_lexicon = dict()

        for blissymbol in all_blissymbols:
            bliss_dict = self.blissymbol_to_dict(blissymbol)
            bliss_dict["BCI-AV"] = int(bci_map[blissymbol.bliss_name])
            eng_words = blissymbol.get_translation("English")
            for eng_word in eng_words:
                bliss_lexicon.setdefault(eng_word, list())
                bliss_lexicon[eng_word].append(bliss_dict)

        self.dump_json(bliss_lexicon, "bliss_lexicon_new")
        return bliss_lexicon

    def refresh_bliss_lexicon(self):
        """
        Refreshes the Blissymbols JSON lexicon to include all new entry in this
        LexiconParser's bliss_dict.

        :return: None
        """
        lexicon = self.translator.find_bliss_dict("English")
        lexicon = {entry: list(filter(lambda x: x is not None and len(x) > 0,
                                 [self.blissymbol_to_dict(bliss) for bliss in lexicon[entry]]))
                   for entry in lexicon}
        self.dump_json(lexicon, "bliss_lexicon")

    # BLISS UNICODE
    # --------------
    def load_bliss_unicode(self):
        """
        Returns a unicode-to-Blissymbol name dictionary.
        ~
        Unicode conforms to suggestions here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict(str, str), where...
            key (str) - unicode for Blissymbol
            val (str) - Blissymbol's name
        """
        return self.load_json("bliss_unicode")

    def fresh_bliss_unicode(self):
        """
        Returns a fresh text_image-to-unicode dictionary for translating
        words to Blissymbols and dumps it to bliss_unicode JSON.

        :return: dict, where...
            key (str) - unicode for a Blissymbol
            val (List[str]) - Blisswords for this unicode
        """
        blisswords = self.load_json("raw_bliss_unicode")
        lexicon = self.load_bliss_lexicon()
        uni = int("3200", base=16)
        unicodes = dict()

        for word in blisswords:
            word = word.lower()
            lookup_word = self.translator.remove_parens(word)
            if lookup_word not in lexicon:
                lookup_word = lookup_word.title()
            hex_uni = "U+" + hex(uni)[-4:]
            unicodes.setdefault(hex_uni, None)

            if lookup_word in lexicon:
                bliss_dicts = lexicon[lookup_word]
                bliss_names = list()
                underscored_word = self.translator.underscore(word)
                for bd in bliss_dicts:
                    name = bd['name']
                    if name[-5:] == "-(to)":
                        continue
                    elif len(bliss_dicts) == 1:
                        bliss_names.append(name)
                    elif word == name:
                        bliss_names.append(name)
                    else:
                        deriv = bd['derivation']
                        is_atom = deriv[-9:] == "Character"
                        is_char = "Character" in deriv or "[modifi" in deriv

                        if is_atom and underscored_word in name:
                            bliss_names.append(name)
                        elif is_char and underscored_word in name:
                            bliss_names.append(name)
                        else:
                            bliss_name = name.split(",")[0]
                            if underscored_word == bliss_name:
                                bliss_names.append(name)

                if len(bliss_names) != 0:
                    unicodes[hex_uni] = bliss_names[0]
            uni += 1

        self.dump_json(unicodes, "bliss_unicode")
        return unicodes

    def refresh_bliss_unicode(self):
        """
        Overwrites the source unicode-to-Blisswords encoding
        dictionary with this LexiconParser's BlissTranslator's
        bliss_unicode dict.

        :return: None
        """
        unicodes = self.translator.init_bliss_unicode()
        self.dump_json(unicodes, "bliss_unicode")

    # BLISSNET
    # --------------
    def load_blissnet(self):
        """
        Returns a unicode-to-Synset dictionary from Blissymbols
        to PWN Synsets, where Synsets are loaded as strings.
        ~
        Unicode conforms to suggestions here:
        http://std.dkuug.dk/JTC1/SC2/WG2/docs/n1866.pdf

        :return: dict, where...
            key (str) - unicode for a Blissymbol
            val (List[str]) - Princeton synsets for given unicode
        """
        return self.load_json("blissnet")

    def init_blissnet(self, reverse=False):
        """
        Returns a Blissymbol-to-Synset dictionary for mapping
        Blissymbols to Wordnet.
        ~
        If reverse is True, output dict is Synset-to-Blissymbols instead.

        :param reverse: bool, whether to switch keys & values
        :return: dict, where...
            key (Blissymbol) - Blissymbol with synsets
            val (List[Synset]) - Princeton synsets for given Blissymbol
        """
        blissnet = self.load_blissnet()
        blissnet = {self.translator.unicode_to_blissymbol(uni):
                        self.translator.strs_synsets(blissnet[uni]) for uni in blissnet}
        if reverse:
            rev_blissnet = dict()

            for blissymbol in blissnet:
                synsets = blissnet[blissymbol]
                for synset in synsets:
                    rev_blissnet.setdefault(synset, list())
                    rev_blissnet[synset].append(blissymbol)

            return rev_blissnet
        else:
            return blissnet

    def fresh_blissnet(self):
        """
        Returns a new Blissymbol-to-Synset dict and overwrites
        blissnet.json with it.

        :return: dict, where...
            key (Blissymbol) - Blissymbol with Synsets
            val (List[Synset]) - Synsets for given Blissymbol
        """
        all_blissymbols = self.init_blissymbols()
        blissnet = dict()

        for blissymbol in all_blissymbols:
            synsets = self.translator.blissymbol_to_synsets(blissymbol)
            print(blissymbol, "has synsets:\n\t", synsets, "\n")
            if len(synsets) != 0:
                '''
                if len(synsets) >= 3:
                    valids = input("That's " + str(len(synsets)) + " synsets.  "
                                       "Write the indices of the best ones, or N if none exist.\n")
                    if len(valids) != 0:
                        if valids == "N":
                            continue
                        valids = eval("[" + valids.replace(" ", ",") + "]")
                        synsets = [synsets[i-1] for i in valids]
                '''
                blissnet.setdefault(blissymbol, synsets[:3])

        self.write_blissnet(blissnet)
        return blissnet

    def write_blissnet(self, blissnet):
        json_blissnet = dict()

        for blissymbol in blissnet:
            synsets = blissnet[blissymbol]
            json_blissnet[blissymbol.unicode] = [s.name() for s in synsets]

        self.dump_json(json_blissnet, "blissnet")

    # BLISSWORDNET
    # ------------
    def load_blisswordnet(self):
        """
        Returns a dictionary from Blissymbol names to their
        corresponding PWN Synsets.
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict, where...
            key (str) - name for a Blissymbol
            val (List[str]) - Princeton synsets for given unicode
        """
        return self.load_json("fresh_blisswordnet")

    def fresh_blisswordnet(self):
        """
        Returns a dict mapping Blissymbol names to
        a list of WordNet synsets and dumps it to
        fresh_blisswordnet JSON.
        ~
        N.B. Blissymbol names are a comma-separated string of
             1 or more English words.

        :return: dict(str, list), where...
            key (str) - name for Blissymbol
            val (List[str]) - synsets for Blissymbol
        """
        blissnet = self.load_blissnet()
        bliss_unis = self.load_bliss_unicode()
        blisswordnet = dict()

        for uni in blissnet:
            bliss_word = bliss_unis[uni]
            blisswordnet[bliss_word] = blissnet[uni]

        self.dump_json(blisswordnet, 'fresh_blisswordnet')
        return blisswordnet

    # WN30-WN31 MAP
    # -------------
    def load_wn30_map_wn31(self):
        """
        Returns a dictionary mapping from WordNet 3.0 to 3.1.
        ~
        Map only contains entry where 3.0 differs from 3.1.

        :return: dict(str, str), where...
            key (str) - WordNet 3.0 synset ID
            val (str) - WordNet 3.1 synset ID
        """
        return self.load_json("wn30map31")

    def fresh_wn30_map_31(self):
        """
        Returns a dict mapping WordNet 3.0 synset ID numbers to
        WordNet 3.1 synset ID numbers, where they differ.
        ~
        Dumps the 3.0-to-3.1 mapping to wn30map31 JSON.
        ~
        N.B. Synset IDs are 8-digit strings of integers.
             Stored as strings since many are preceded by zeroes.

        :return: dict(str, str), where...
            key (str) - WordNet 3.0 synset ID number
            val (str) - corresponding WordNet 3.1 synset ID number
        """
        wn30_wn31_txt = self.load_wn30_map_wn31()
        wn30_wn31_map = dict()

        for line in wn30_wn31_txt.readlines():
            if line[0] == "#":  # skip header
                continue
            else:
                line = line.replace("\n", "")
                pos, wn30, wn31 = line.split("\t")
                if wn30 != wn31:  # if wn 3.0 & 3.1 ids differ,
                    wn30_wn31_map[wn30] = wn31  # add to dict

        self.dump_json(wn30_wn31_map, "wn30map31")
        return wn30_wn31_map

    def convert_wn30_31(self, sensekey, wn30_in=True):
        """
        Translates this sensekey from WordNet 3.0 to 3.1
        if wn30_in is True.  Otherwise, translates from WN
        3.1 to 3.0.

        :param sensekey: str, 8-digit sense-key for a synset in WN 3.0/3.1
        :param wn30_in: bool, whether to translate from 3.0 to 3.1
        :return: str, 8-digit sense-key for a synset in WN 3.1/3.0
        """
        wn_map = self.load_wn30_map_wn31()
        if not wn30_in:
            wn_map = {v: k for k,v in wn_map.items()}
        return wn_map.get(sensekey, None)


    # MULTILINGUAL
    # ============
    def parse_lexicon(self, language):
        """
        Parses plaintext file for given language.
        Returns a dict with all words in lexicon
        as keys and corresponding lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes first word on every line is the lemma form
        of all subsequent words on the same line.
        ~
        N.B. The same lemma value will often belong to multiple keys.

        e.g. "kota, kocie, kot" -> {"kota":"kota", "kocie":"kota", "kot":"kota"}

        :param language: str, language of .txt file for lexicon
        :return: dict, where...
            key (str) - inflected form of a word
            val (str) - lemma form of inflected word
        """
        filename = "/resources/lexica/" + language + ".txt"
        lexicon = None

        with open(self.FILE_PATH + filename, "r", encoding='utf-8') as lex:
            if language == "Polish":
                lexicon = self.parse_polish_lexicon(lex)
            elif language == "French":
                lexicon = self.parse_french_lexicon(lex)

        return lexicon

    def parse_french_lexicon(self, lex):
        """
        Parses plaintext file for French lexicon with
        given filename.  Returns a dict with all lexemes
        in French lexicon as keys and corresponding
        lemma forms as values.
        ~
        Each new lexical entry should be separated by "\n",
        while inflected forms should be separated by ",".
        ~
        Assumes second word on every line is the lemma form
        of previous words on the same line.
        ~
        N.B. The same lemma value will often belong to
        multiple lexemes.

        e.g. "grande, grand" -> {"grande":"grand", "grand":"grand"}

        :param lex: List[str], entry in French lexicon
        :return: dict, where...
            key (str) - any lexical form of a word
            val (str) - lemmatized form of lemma
        """
        lexicon = {}

        for line in lex:
            line = line.decode("utf-8")
            line = line.strip("\n")
            if line[-1] == "=":
                lemma = line[:-2]
                lexicon[lemma] = lemma
            else:
                entry = line.split("\t")
                lemma = entry[1]
                lexeme = entry[0]
                lexicon[lexeme] = lemma

        return lexicon

    def parse_polish_lexicon(self, lex):
        """
        Parses plaintext file for Polish lexicon with
        given filename.  Returns a dict with all lexemes
        in Polish lexicon as keys and corresponding
        lemma forms as values.
        ~
        Inflected forms in each lexical entry should be
        separated by ",".
        ~
        Assumes first word on every line is the lemma form
        of following words on the same line.
        ~
        N.B. The same lemma value will often belong to
        multiple lexemes.

        e.g. "kota, kot, kocie" -> {"kota":"kota", "kot":"kota", "kocie":"kota"}

        :param lex: List[str], entry in Polish lexicon
        :return: dict, where...
            key (str) - any lexical form of a word
            val (str) - lemmatized form of lemma
        """
        lexicon = {}

        for line in lex:
            line = str(line)
            line = line.strip("\n")
            line = line.strip("\r")
            inflexions = line.split(",")
            lemma = inflexions[0]

            for inflexion in inflexions:
                lexicon[inflexion.strip()] = lemma

        return lexicon


    # WORDNET
    # =======
    def bliss_wordnet_mapping(self, bliss_keys=True, bliss_attr='gloss', wn_attr='name',
                              wn_30=True, title='bliss_wordnet_mapping'):
        """
        Returns a dict mapping between Blissymbols and WordNet and dumps it to
        /resources/data/[title].json.

        :param bliss_keys: bool, True if Blissymbols should be keys in map, False if values
        :param bliss_attr: str, which
            MUST be one of:
                'gloss' - Blissymbol's gloss, AKA its "bliss-name" (e.g. "cat,feline_(animal),felid")
                'unicode' - Blissymbol's unicode representation (e.g. "U+349b")
                'number' - Blissymbol's BCI-AV# (e.g. 12383)
        :param wn_attr: str,
            MUST be one of:
                'sense-key' - Synset's gloss, AKA its "bliss-name" (e.g. "cat,feline_(animal),felid")
                'name' - Synset's first lemma, pos, and sense number (e.g. "cat.n.01")
                'lemma_names' - Synset's lemmas' names (e.g. ["cat", "feline", "felid"])
        :param wn_30: bool,
        :param title: str, name of JSON file to dump mapping to
        :return: dict(X, Y), Blissymbols mapped to Synsets OR Synsets mapped to Blissymbols
        """
        all_blissymbols = self.init_blissymbols()
        bliss_wn_map = dict()

        def transform_bliss(bs):
            if bliss_attr == 'unicode':
                return bs.unicode
            elif bliss_attr == 'number':
                return bs.bci_num
            else:
                try:
                    return getattr(bs, bliss_attr)
                except AttributeError:
                    return bs.bliss_name

        def transform_synset(s):
            if wn_attr == 'sense-key':
                self.translator.synset_id()
                synset_30 = str(s.offset()).zfill(8)
                if not wn_30:
                    return self.convert_wn30_31(synset_30, wn30_in=True)
                else:
                    return synset_30
            else:
                try:
                    return getattr(s, wn_attr)
                except AttributeError:
                    return s.name()

        for blissymbol in all_blissymbols:
            synsets = self.translator.blissymbol_to_synsets(blissymbol)
            if len(synsets) != 0:
                trans_bliss = transform_bliss(blissymbol)
                trans_synsets = [transform_synset(ss) for ss in synsets]
                if bliss_keys:
                    bliss_wn_map[trans_bliss] = trans_synsets
                else:
                    for trans_synset in trans_synsets:
                        bliss_wn_map.setdefault(trans_synset, list()).append(trans_bliss)

        self.dump_json(bliss_wn_map, filename=title)
        return bliss_wn_map

    def blissnum_sensekey_mapping(self, wn_version=3.0):
        """
        Returns a dict mapping Blissymbol BCI_AV# to WordNet 3.0 or 3.1 synset sense-key,
        and dumps it to bci_blissnet_3.0/3.1 JSON.
        ~
        Synset id # corresponds to WordNet version specified in wn_version (either 3.0 or 3.1).
        ~
        N.B. WordNet sense-keys are 8-digit strings (can start with 0) of numbers

        :param wn_version: float, version of WordNet to use (3.0 or 3.1)
        :return: dict(int, list), where...
            key (int) - BCI_AV# for a Blissymbol
            val (List[str]) - synset sense-key for this Blissymbol
        """
        bci_blissnet = self.load_bci_blissnet()
        blissnet_30 = dict()

        for bci in bci_blissnet:
            synset_strs = bci_blissnet[bci]
            synset_30_ids = [str(self.translator.str_synset_id(s))[1:] for s in synset_strs]
            blissnet_30[int(bci)] = synset_30_ids

        if wn_version == 3.0:
            self.dump_json(blissnet_30, "bci_blissnet_3.0")
            return blissnet_30
        elif wn_version == 3.1:
            wn30_map_31 = self.load_wn30_map_wn31()
            blissnet_31 = {int(bci): [wn30_map_31.get(id30, id30) for id30 in ids30]
                           for bci, ids30 in blissnet_30.items()}
            self.dump_json(blissnet_31, "bci_blissnet_3.1")
            return blissnet_31

    def synset_sensekey_mapping(self):
        """
        Maps WordNet 3.0 synsets to sense-keys.

        :return: dict(str, str)
        """
        synsets = self.translator.all_synsets()
        mapping = dict()

        for synset in synsets:
            str_synset = synset.name()
            sense_key = str(synset.offset()).zfill(8)
            mapping[str_synset] = sense_key

        self.dump_json(mapping, "synset_sense-key_mapping")
        return mapping

    def bliss_dict_to_wordnet(self, bliss_dict):
        """
        Returns a dictionary of Blissymbol word keys and synsets
        from this bliss_dict.
        ~
        N.B. Output will be most accurate with a multilingual
        Bliss dictionary.

        :param bliss_dict: dict, where...
            key (str) - Blissymbol word in English
            val (List[Blissymbol]) - corresponding Blissymbols with
                translations in all languages
        :return: dict, where...
            key (str) - Blissymbol's unicode representation
            val (List[Synset]) - synsets associated with Blissymbol
        """
        wordnet = {}

        for word in bliss_dict:
            for blissymbol in bliss_dict[word]:
                uni = blissymbol.unicode
                synsets = blissymbol.synsets
                wordnet[uni] = synsets

        return wordnet

    def get_tab_file(self, lang_code):
        """
        Retrieves a multilingual WordNet tab file for the
        given language.
        ~
        If no such tab file exists, returns None.

        :param lang_code: str, 3-character ISO language code
        :return: tab file, WordNet file for given language
        """
        try:
            tab_path = "/resources/omw_tabs/" + "wn-cldr-" + lang_code + ".tab"
            return open(self.FILE_PATH + tab_path, encoding='utf-8')
        except IOError:
            return None

    # BLISS MANIPULATION
    # ==================
    def make_blissymbol(self):
        """
        Prompts user for new Blissymbol entry information.

        :return: Blissymbol, represents 1 Bliss lexical entry
        """
        bliss_name = input("What do you call your new Blissymbol? ")
        bliss_filename = bliss_name + ".png"
        print("Which part of speech is this? ")
        code = Blissymbol.POS_COLOUR_CODE

        for pos in code:
            print(str(pos) + ":\t" + str(code[pos]))

        pos = input("")
        print("Which atomic Blissymbols is this made of? ")
        derivations = input("Separate them by commas: ")
        derivations = derivations.split(",")
        derivs = ""
        derivs += "("

        for idx in range(0, len(derivations)):
            derivs += derivations[idx]
            if idx == len(derivations)-1:
                derivs += ")"
            else:
                derivs += " + "

        translations = {}

        for language in self.BLISS_LANGS:
            print("What is/are the translation(s) in " + language + "? ")

            try:
                translation = input("Separate by commas if necessary: ")
            except SyntaxError:
                continue
            else:
                translation = translation

            translations.setdefault(language, [])
            translations[language].append(translation)

        blissymbol = Blissymbol(img_filename=bliss_filename, pos=pos, derivation=derivs,
                                translations=translations, translator=self.translator, num=0)
        return blissymbol

    def blissymbol_entry(self, name, pos, num, derivation, translations):
        """
        Returns a dict of these inputs as a Blissymbol's initializing parameters,
        i.e., its name, pos, BCI-AV#, derivation, and translations.

        :param name: str, Blissymbol's name
        :param pos: str, Blissymbol's part-of-speech
        :param num: str, Blissymbol's BCI-AV#
        :param derivation: str, Blissymbol's derivation
        :param translations: dict(str, list), languages and translations of Blissymbol
        :return: dict, where...
            key (str) - name of Blissymbol field
            val (X) - corresponding value
        """
        return {"name": name,
                "pos": pos,
                "BCI-AV": num,
                "derivation": derivation,
                "translations": translations}

    def blissymbol_to_dict(self, blissymbol):
        """
        Returns a dict of this Blissymbol's initializing
        parameters, i.e. its name, pos, derivation, and translations.

        :param blissymbol: Blissymbol, symbol to turn to tuple
        :return: dict, where...
            key (str) - name of Blissymbol field
            val (X) - corresponding value
        """
        return {"name": blissymbol.bliss_name,
                "pos": blissymbol.get_pos(),
                "BCI-AV": blissymbol.bci_num,
                "derivation": blissymbol.derivation,
                "translations": blissymbol.translations}

    def dict_to_blissymbol(self, d):
        """
        Returns this dict, d, as a Blissymbol.
        ~
        This method assumes d has keys named
        "name", "pos", "derivation", and "translations",
        and initializes a Blissymbol with its fields corresponding
        to these keys.

        :param d: dict, dictionary to turn into Blissymbol
        :return: Blissymbol, d as a Blissymbol
        """
        name = d["name"]

        if name[-7:-1] != "ercase" and name[-4:-1] != "OLD":
            blissymbol = Blissymbol(name+".png",
                                    d["pos"],
                                    d.get("derivation", ""),
                                    d.get("translations", dict()),
                                    self.translator,
                                    num=d["BCI-AV"])
            return blissymbol

    # XLSX ENTRIES
    # ============
    def load_bliss_xlsx(self):
        """
        Loads a new Blissymbols lexicon from the
        "universal bliss lexicon.xlsx" file.

        :return: List[dict], list of dicts for all Blissymbols
        """
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        bliss_dicts = list()

        def split_translation(entry): return [self.translator.deunderscore(e)
                                              for e in entry.split(",") if len(e) != 0]

        for cells in sheet.iter_rows(min_row=2):
            eng_cell = cells[1]
            eng_word = eng_cell.value
            eng_words = split_translation(eng_word)

            pos_cell = cells[2]
            pos_colour = pos_cell.value

            derivs_cell = cells[3]
            derivation = derivs_cell.value

            translations_dict = {"English": eng_words}

            for i in range(5, len(self.LEXICON_COLS), 2):
                lang_cell = cells[i]
                cell_language = self.LEXICON_COLS[i]
                translation = lang_cell.value
                translations = split_translation(translation) if translation is not None else list()
                translation_dict = {cell_language: translations}
                translations_dict.update(translation_dict)

            bliss_poses = Blissymbol.find_colour_pos(pos_colour)
            pos = self.translator.token_pos(eng_words[0])
            pos_lst = [pos]
            poses = list(set(bliss_poses).intersection(pos_lst))
            if len(poses) == 0:
                poses = bliss_poses
            pos = poses[0]
            entry = self.blissymbol_entry(eng_word, pos, derivation, translations_dict)
            bci_num = cells[0].value
            entry["BCI-AV"] = bci_num
            bliss_dicts.append(entry)

        self.dump_json(bliss_dicts, "all_blissymbols")
        return bliss_dicts

    def blissymbol_to_xlsx_entry(self, blissymbol):
        """
        Converts this Blissymbol to a list of
        information constituting a Bliss lexicon entry.
        ~
        Must be in order of LEXICON_COLS.

        :param blissymbol: Blissymbol, symbol to convert to entry
        :return: List[str], Bliss lexicon entry
        """
        row = []
        bci_col = self.LEXICON_COLS[0]
        pos_col = self.LEXICON_COLS[2]
        deriv_col = self.LEXICON_COLS[3]
        uni = blissymbol.unicode
        uni = uni[2:]

        for col in self.LEXICON_COLS:
            if col == bci_col:
                row.append("C+" + uni)
            elif col == pos_col:
                row.append(blissymbol.pos_to_int())
            elif col == deriv_col:
                row.append(blissymbol.derivation)
            else:
                translations = blissymbol.get_translation(col)
                translations = ",".join(translations)
                #translations = self.translator.deunicodize(translations)
                row.append(translations)

        return row

    def append_bliss_xlsx_entry(self, blissymbol=None):
        """
        Adds data to bliss lexicon file for later use.
        ~
        If no Blissymbol is provided, prompts you to create your
        own custom symbol.

        :param blissymbol: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        if not blissymbol:
            blissymbol = self.make_blissymbol()

        print("making new Blissymbol entry for", str(blissymbol))
        bliss_entry = self.blissymbol_to_xlsx_entry(blissymbol)

        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        sheet.append(bliss_entry)

        book.save(self.LEXICON_PATH)

    def extend_bliss_xlsx_entry(self, blissymbol):
        """
        Extends data in bliss lexicon file for later use.

        :param bliss_entry: Blissymbol, entry to add to Bliss lexicon
        :return: None
        """
        print("extending entry for", blissymbol.get_bliss_name())
        book = load_workbook(self.LEXICON_PATH)
        sheet = book.worksheets[0]
        start_row = sheet.min_row
        end_row = sheet.max_row

        for row_idx in range(start_row, end_row):
            cell = sheet.cell(row=row_idx, column=2)
            eng_word = cell.value
            eng_words = eng_word.split(",")
            for synonym in eng_words:
                synonym = synonym.strip()
                if synonym == blissymbol.get_bliss_name():
                    print("found English Blissymbol synonym:", eng_word)
                    translations = blissymbol.translations
                    for language in translations:
                        defns = translations[language]
                        if len(defns) != 0:
                            lang_idx = self.LEXICON_COLS.index(language)
                            edit = sheet.cell(row=row_idx, column=lang_idx+1)

                            if edit.value is None:
                                edit.value = ""
                            else:
                                edit.value = str(edit.value)
                            values = []
                            for defn in defns:
                                values = edit.value.split(",")
                                if defn not in values:
                                    values.append(defn)
                            values = self.translator.remove_duplicates(values)
                            edit.value = ",".join(values)
                            print("edited", language, "to be", edit.value)
                    break

        book.save(self.LEXICON_PATH)

